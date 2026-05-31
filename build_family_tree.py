"""Generate an Excel workbook from the hand-drawn Kalaimalai family tree.

Readings marked with "(?)" are uncertain transcriptions from the handwriting
and should be verified by the family. Use the "Verify?" column to track checks.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
GEN_FILLS = {
    0: PatternFill("solid", fgColor="C6E0B4"),  # ancestors
    1: PatternFill("solid", fgColor="FFE699"),  # trunk (only son)
    2: PatternFill("solid", fgColor="BDD7EE"),  # children
    3: PatternFill("solid", fgColor="D9D9D9"),  # grandchildren
    4: PatternFill("solid", fgColor="F2F2F2"),  # great-grandchildren
}
UNCERTAIN_FONT = Font(color="C00000")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


# ===========================================================================
# Sheet 1: People table (flat, structured)
# ===========================================================================
ws = wb.active
ws.title = "People"

headers = ["Branch", "Generation", "Person", "Spouse", "Parent(s) / Belongs to", "Verify?", "Notes"]
ws.append(headers)
style_header(ws, len(headers))

# Each row: (branch, generation, person, spouse, parent, notes)
# "(?)" marks an uncertain reading of the handwriting.
rows = [
    # --- Ancestors (from the legend at the bottom of the page) ---------------
    ("Ancestors", 0, "Muddukrishnaiah(?)", "Seshamma", "—",
     "Legend ③: 'Pre Pre Praternala(?) Muddukrishnaiah / Seshamma' — earliest named couple."),
    ("Ancestors", 0, "Veeraraghavaiah(?)", "Konakavalli Thayar(?)", "Muddukrishnaiah & Seshamma",
     "Legend ②: 'Pre Pillumeala(?) Veeraraghavaiah / Konakavalli Thayar'."),
    ("Ancestors", 0, "Pritha Krathi(?)", "Jaganmasi(?) / Annabavula(?)", "Veeraraghavaiah & Konakavalli Thayar",
     "Legend ①: 'Pritha Krathi(?) / Jaganmasi(?) / Annabavula(?)' — hard to read."),

    # --- Trunk: the central couple ('Only Son') -----------------------------
    ("TRUNK", 1, "Sri Kalaimalai Veeraraghavaiah", "Smt. Kalaimalai Konakavalli Thayaramma",
     "Kalaimalai Muddukrishnaiah & Smt. Seshamma",
     "Central trunk, labelled 'Only Son'. Parents named bottom-left of the drawing."),

    # --- Branch 1 (red-circled ①, lower centre) ------------------------------
    ("Branch 1", 2, "Jagan(?)", "Allour(?) / Prabhavalli(?)", "Trunk couple",
     "Branch ①. 'Jagan + allour / Prabhavalli' — names uncertain."),
    ("Branch 1", 3, "Lakshman(?) + kumar(?)", "", "Jagan(?)",
     "Children listed: 'Lakshman+kumar', 'Nitin', 'Jagan'."),
    ("Branch 1", 3, "Nitin(?)", "", "Jagan(?)", ""),
    ("Branch 1", 3, "Jagan(?)", "", "Jagan(?)", ""),

    # --- Branch 2 (red-circled ②, lower right) -------------------------------
    ("Branch 2", 2, "Vij(?) / Malti(?)", "Thiago(?)", "Trunk couple",
     "Branch ②. Top label 'Vij Malti Thiago' — uncertain."),
    ("Branch 2", 3, "Abhena(?)", "Ramu(?)", "Branch 2", "'Abhena + Ramu'."),
    ("Branch 2", 3, "Vaibhav(?)", "Surya(?)", "Branch 2", "'Vaibhav + Surya'."),
    ("Branch 2", 3, "Surya(?)", "Balaji(?)", "Branch 2", "'Surya + Balaji'."),
    ("Branch 2", 3, "Pradeep(?) + kumar(?)", "Neha(?)", "Branch 2",
     "'Pradeep+kumar', children 'Neha', 'Tekil(?)'."),
    ("Branch 2", 4, "Neha(?)", "", "Pradeep(?)", ""),
    ("Branch 2", 4, "Tekil(?)", "", "Pradeep(?)", ""),
    ("Branch 2", 3, "Shravan(?)", "Shreyas(?)", "Branch 2", "'Shravan + Shreyas'."),

    # --- Branch 3 (red-circled ③, middle right) ------------------------------
    ("Branch 3", 2, "Krishnamoorthi(?)", "Chithra(?)", "Trunk couple",
     "Branch ③. 'Krishnamoorthi ... adirudine syndrome(?)' — note appears to be a medical annotation."),
    ("Branch 3", 3, "Madhubabu(?)", "Lavni(?)", "Branch 3",
     "'Madhubabu / Lavni', children 'Bumika', 'Kranesa'."),
    ("Branch 3", 4, "Bumika(?)", "", "Madhubabu(?)", ""),
    ("Branch 3", 4, "Kranesa(?)", "", "Madhubabu(?)", ""),
    ("Branch 3", 3, "Usha(?)", "Ravi(?)", "Branch 3",
     "'Usha + Ravi', children 'Aditya', 'Tejal'."),
    ("Branch 3", 4, "Aditya(?)", "", "Usha(?) & Ravi(?)", ""),
    ("Branch 3", 4, "Tejal(?)", "", "Usha(?) & Ravi(?)", ""),

    # --- Branch 4 (red-circled ④, left side) ---------------------------------
    ("Branch 4", 2, "Devendra(?)", "", "Trunk couple",
     "Branch ④ (left side of tree). Several left-hand labels are illegible."),
    ("Branch 4", 3, "Padmavati(?)", "Rangaswami(?) / Rasi(?)", "Branch 4",
     "'Padmavati + Rangaswami(?)'."),
    ("Branch 4", 3, "Harini(?)", "Ravindravani(?)", "Branch 4",
     "'Harini + Ravindravani', children 'Magta(?)', 'Tejodha(?)'."),
    ("Branch 4", 4, "Magta(?)", "", "Harini(?)", ""),
    ("Branch 4", 4, "Tejodha(?)", "", "Harini(?)", ""),
    ("Branch 4", 3, "[illegible]", "", "Branch 4",
     "One or more left-hand labels could not be read from the photo."),

    # --- Branch 5 (red-circled ⑤, top right) ---------------------------------
    ("Branch 5", 2, "Kodandaramaiah(?)", "Rajeswari(?)", "Trunk couple",
     "Branch ⑤ (top right). 'Kodandaramaiah + Rajeswari'."),
    ("Branch 5", 3, "Deena(?)", "Vasishwari(?)", "Branch 5",
     "'Deena + Vasishwari', child 'Vignesh'."),
    ("Branch 5", 4, "Vignesh(?)", "", "Deena(?)", ""),
    ("Branch 5", 3, "Manya(?)", "Bhaskar(?)", "Branch 5",
     "'Manya / Humsavini(?) / Bhaskar'."),
    ("Branch 5", 4, "Humsavini(?)", "", "Manya(?) & Bhaskar(?)", ""),
    ("Branch 5", 3, "Hemat(?) / Salwan(?)", "Aswini(?)", "Branch 5",
     "Top cluster: 'Hemat / Salwan / Aswini / Gotali(?) / Aswini' — very hard to read."),
    ("Branch 5", 4, "Gotali(?)", "", "Hemat(?)", ""),
    ("Branch 5", 4, "Aswini(?)", "", "Hemat(?)", ""),
]

for r in rows:
    ws.append(r)

# Apply per-row styling
for ridx in range(2, ws.max_row + 1):
    gen = ws.cell(row=ridx, column=2).value
    fill = GEN_FILLS.get(gen)
    for cidx in range(1, len(headers) + 1):
        cell = ws.cell(row=ridx, column=cidx)
        cell.border = BORDER
        cell.alignment = WRAP
        if fill:
            cell.fill = fill
        # highlight uncertain readings in red text
        val = cell.value
        if isinstance(val, str) and ("(?)" in val or "illegible" in val):
            cell.font = UNCERTAIN_FONT

widths = [12, 11, 26, 26, 30, 9, 52]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 22

# ===========================================================================
# Sheet 2: Tree outline (indented, mirrors the drawing's hierarchy)
# ===========================================================================
ws2 = wb.create_sheet("Tree Outline")
ws2.append(["Gen 0 (ancestors)", "Gen 1 (trunk)", "Gen 2 (children)",
            "Gen 3 (grandchildren)", "Gen 4 (great-grandchildren)"])
style_header(ws2, 5)

outline = [
    ["Muddukrishnaiah(?) = Seshamma", "", "", "", ""],
    ["", "Sri Kalaimalai Veeraraghavaiah = Smt. Konakavalli Thayaramma (Only Son)", "", "", ""],
    ["", "", "① Jagan(?) = Prabhavalli(?)", "Lakshman+kumar(?)", ""],
    ["", "", "", "Nitin(?)", ""],
    ["", "", "", "Jagan(?)", ""],
    ["", "", "② Vij(?)/Malti(?) = Thiago(?)", "Abhena(?) = Ramu(?)", ""],
    ["", "", "", "Vaibhav(?) = Surya(?)", ""],
    ["", "", "", "Surya(?) = Balaji(?)", ""],
    ["", "", "", "Pradeep+kumar(?)", "Neha(?)"],
    ["", "", "", "", "Tekil(?)"],
    ["", "", "", "Shravan(?) = Shreyas(?)", ""],
    ["", "", "③ Krishnamoorthi(?) = Chithra(?)", "Madhubabu(?) = Lavni(?)", "Bumika(?)"],
    ["", "", "", "", "Kranesa(?)"],
    ["", "", "", "Usha(?) = Ravi(?)", "Aditya(?)"],
    ["", "", "", "", "Tejal(?)"],
    ["", "", "④ Devendra(?)", "Padmavati(?) = Rangaswami(?)", ""],
    ["", "", "", "Harini(?) = Ravindravani(?)", "Magta(?)"],
    ["", "", "", "", "Tejodha(?)"],
    ["", "", "", "[illegible left-side labels]", ""],
    ["", "", "⑤ Kodandaramaiah(?) = Rajeswari(?)", "Deena(?) = Vasishwari(?)", "Vignesh(?)"],
    ["", "", "", "Manya(?) = Bhaskar(?)", "Humsavini(?)"],
    ["", "", "", "Hemat(?)/Salwan(?) = Aswini(?)", "Gotali(?)"],
    ["", "", "", "", "Aswini(?)"],
]
for row in outline:
    ws2.append(row)

for ridx in range(2, ws2.max_row + 1):
    for cidx in range(1, 6):
        cell = ws2.cell(row=ridx, column=cidx)
        cell.border = BORDER
        cell.alignment = WRAP
        if cidx in GEN_FILLS and cell.value:
            cell.fill = GEN_FILLS[cidx - 1]
        if isinstance(cell.value, str) and ("(?)" in cell.value or "illegible" in cell.value):
            cell.font = UNCERTAIN_FONT

for i in range(1, 6):
    ws2.column_dimensions[get_column_letter(i)].width = 38
ws2.freeze_panes = "A2"

# ===========================================================================
# Sheet 3: Legend / how to read
# ===========================================================================
ws3 = wb.create_sheet("README")
notes = [
    ["How to read this workbook", ""],
    ["", ""],
    ["Source", "Hand-drawn family tree photo (PHOTO-2022-03-18-20-41-21.jpg)."],
    ["Red text / (?)", "Uncertain reading of the handwriting — PLEASE VERIFY and correct."],
    ["[illegible]", "Could not be read from the photo; needs the original / a clearer scan."],
    ["", ""],
    ["Colour key (generation)", ""],
    ["Green", "Gen 0 — ancestors (from the legend at the bottom of the drawing)."],
    ["Yellow", "Gen 1 — the central trunk couple, labelled 'Only Son'."],
    ["Blue", "Gen 2 — the five main branches ①–⑤ (children)."],
    ["Grey", "Gen 3 — grandchildren and their spouses ('=' / '+' couples)."],
    ["Light grey", "Gen 4 — great-grandchildren."],
    ["", ""],
    ["Conventions", ""],
    ["'=' or '+'", "Marriage / couple."],
    ["①–⑤", "The five red-circled main branches in the drawing."],
    ["", ""],
    ["Trunk couple", "Sri Kalaimalai Veeraraghavaiah  =  Smt. Kalaimalai Konakavalli Thayaramma"],
    ["Trunk parents", "Kalaimalai Muddukrishnaiah  =  Smt. Seshamma"],
]
for r in notes:
    ws3.append(r)
ws3.cell(row=1, column=1).font = Font(bold=True, size=14)
for ridx in range(1, ws3.max_row + 1):
    label = ws3.cell(row=ridx, column=1)
    if label.value and label.value not in ("How to read this workbook", ""):
        label.font = Font(bold=True)
    ws3.cell(row=ridx, column=2).alignment = WRAP
ws3.column_dimensions["A"].width = 24
ws3.column_dimensions["B"].width = 70

out = "/Users/mkalaimalai/github-repos/Kalaimalai/Kalaimalai_Family_Tree.xlsx"
wb.save(out)
print("Saved:", out)
print("Sheets:", wb.sheetnames)
print("People rows:", ws.max_row - 1)
