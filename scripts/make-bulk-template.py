"""Generate a bulk-entry Excel template matching the Genealogy data model.

Sheets:
  - Instructions: how to fill it in
  - People:        one row per person (Key is a local handle used by Relationships)
  - Relationships: one row per edge, referencing People by Key

Run: python3 scripts/make-bulk-template.py
"""

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "docs/family-fabric-bulk-template.xlsx"

# --- enum values (mirrors the domain) ---
GENDER = ["Male", "Female", "Other", "Unknown"]
VISIBILITY = ["Public", "Restricted", "Unlisted"]
BRANCH = ["Maternal", "Paternal"]
REL_TYPE = ["ParentOf", "SpouseOf", "SiblingOf"]
REL_QUALIFIER = ["", "adoption", "step", "remarriage"]

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
REQ_FILL = PatternFill("solid", fgColor="C55A11")  # required cols
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="2F5496")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# ----------------------------------------------------------------------------
# Instructions sheet
# ----------------------------------------------------------------------------
ins = wb.active
ins.title = "Instructions"
ins.sheet_view.showGridLines = False
ins.column_dimensions["A"].width = 110

lines = [
    ("How to fill in this workbook", "title"),
    ("", None),
    ("This template lets you enter your family in bulk. Fill in the two tabs at the", "body"),
    ("bottom: 'People' first, then 'Relationships'.", "body"),
    ("", None),
    ("PEOPLE tab", "h2"),
    ("• One row per person.", "body"),
    ("• Key — a short unique handle you invent (e.g. P1, P2, or 'grandpa'). It is only", "body"),
    ("  used inside this spreadsheet so the Relationships tab can point at people", "body"),
    ("  before the app has assigned real IDs. Every Key must be unique.", "body"),
    ("• Legal Name — REQUIRED. Everything else is optional.", "body"),
    ("• Gender — one of: Male, Female, Other, Unknown.", "body"),
    ("• Visibility — one of: Public, Restricted, Unlisted (defaults to Public).", "body"),
    ("• Branch — Maternal or Paternal.", "body"),
    ("• Birth/Passing dates are split into Year / Month / Day so approximate dates", "body"),
    ("  work: fill only the parts you know (e.g. just Year). Leave blank if unknown.", "body"),
    ("  Passing date cannot be before birth date.", "body"),
    ("", None),
    ("RELATIONSHIPS tab", "h2"),
    ("• One row per relationship.", "body"),
    ("• From Key / To Key — use the Key values from the People tab.", "body"),
    ("• Type:", "body"),
    ("    ParentOf  → From is the PARENT, To is the CHILD (direction matters!)", "body"),
    ("    SpouseOf  → From and To are spouses (order doesn't matter)", "body"),
    ("    SiblingOf → From and To are siblings (order doesn't matter)", "body"),
    ("• Qualifier (optional) — adoption, step, or remarriage. Leave blank for normal.", "body"),
    ("• Example: to record that P1 is the father of P3, put From Key=P1, To Key=P3,", "body"),
    ("  Type=ParentOf. A child with two parents needs two ParentOf rows.", "body"),
    ("", None),
    ("Orange column headers are REQUIRED. Blue headers are optional.", "note"),
    ("Cells with dropdowns will reject values outside the allowed list.", "note"),
]
r = 1
for text, kind in lines:
    c = ins.cell(row=r, column=1, value=text)
    if kind == "title":
        c.font = TITLE_FONT
    elif kind == "h2":
        c.font = Font(bold=True, size=12, color="C55A11")
    elif kind == "note":
        c.font = Font(italic=True, size=10, color="595959")
    else:
        c.font = Font(size=11)
    r += 1


def style_header(ws, headers, required_idx):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = REQ_FILL if (i - 1) in required_idx else HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def add_list_dv(ws, col_idx, values, n_rows=500):
    col = get_column_letter(col_idx)
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(v for v in values if v != "") + '"',
        allow_blank=True,
        showErrorMessage=True,
    )
    dv.error = "Pick a value from the dropdown list."
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{n_rows + 1}")


def add_int_dv(ws, col_idx, lo, hi, n_rows=500):
    col = get_column_letter(col_idx)
    dv = DataValidation(
        type="whole", operator="between", formula1=str(lo), formula2=str(hi),
        allow_blank=True, showErrorMessage=True,
    )
    dv.error = f"Enter a whole number between {lo} and {hi} (or leave blank)."
    dv.errorTitle = "Invalid number"
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{n_rows + 1}")


# ----------------------------------------------------------------------------
# People sheet
# ----------------------------------------------------------------------------
ppl = wb.create_sheet("People")
people_headers = [
    "Key", "Legal Name", "Nickname", "Gender",
    "Birth Year", "Birth Month", "Birth Day",
    "Passing Year", "Passing Month", "Passing Day",
    "Birthplace", "Current Location", "Bio", "Visibility", "Branch",
]
people_required = {0, 1}  # Key, Legal Name
style_header(ppl, people_headers, people_required)

widths = [10, 26, 16, 12, 11, 12, 10, 13, 14, 12, 22, 22, 40, 13, 12]
for i, w in enumerate(widths, start=1):
    ppl.column_dimensions[get_column_letter(i)].width = w

add_list_dv(ppl, 4, GENDER)            # Gender
add_int_dv(ppl, 5, 1800, 2100)         # Birth Year
add_int_dv(ppl, 6, 1, 12)              # Birth Month
add_int_dv(ppl, 7, 1, 31)              # Birth Day
add_int_dv(ppl, 8, 1800, 2100)         # Passing Year
add_int_dv(ppl, 9, 1, 12)              # Passing Month
add_int_dv(ppl, 10, 1, 31)             # Passing Day
add_list_dv(ppl, 14, VISIBILITY)       # Visibility
add_list_dv(ppl, 15, BRANCH)           # Branch

# sample rows (greyed hint) — based on the seeded family
samples = [
    ["P1", "Krishna Murthy K V", "Appa", "Male", 1940, "", "", 2018, "", "", "Tirupati", "", "Patriarch; schoolteacher.", "Public", "Paternal"],
    ["P2", "Indrani V R", "Amma", "Female", 1945, "", "", "", "", "", "Madurai", "", "", "Public", "Paternal"],
    ["P3", "Madhu Kalaimalai", "", "Male", 1969, 9, 10, "", "", "", "", "", "Civil engineer; Software Architect.", "Public", "Paternal"],
]
for ri, row in enumerate(samples, start=2):
    for ci, val in enumerate(row, start=1):
        c = ppl.cell(row=ri, column=ci, value=val if val != "" else None)
        c.font = Font(italic=True, color="808080")
        c.border = BORDER

# ----------------------------------------------------------------------------
# Relationships sheet
# ----------------------------------------------------------------------------
rel = wb.create_sheet("Relationships")
rel_headers = ["From Key", "To Key", "Type", "Qualifier"]
rel_required = {0, 1, 2}
style_header(rel, rel_headers, rel_required)
for i, w in enumerate([14, 14, 14, 14], start=1):
    rel.column_dimensions[get_column_letter(i)].width = w

add_list_dv(rel, 3, REL_TYPE)        # Type
add_list_dv(rel, 4, REL_QUALIFIER)   # Qualifier

rel_samples = [
    ["P1", "P2", "SpouseOf", ""],
    ["P1", "P3", "ParentOf", ""],
    ["P2", "P3", "ParentOf", ""],
]
for ri, row in enumerate(rel_samples, start=2):
    for ci, val in enumerate(row, start=1):
        c = rel.cell(row=ri, column=ci, value=val if val != "" else None)
        c.font = Font(italic=True, color="808080")
        c.border = BORDER

# add a helper note under the type column header area
rel.cell(row=1, column=6, value="ParentOf: From=parent, To=child").font = Font(italic=True, size=9, color="808080")

wb.save(OUT)
print("wrote", OUT)
